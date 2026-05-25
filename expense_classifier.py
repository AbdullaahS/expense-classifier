import anthropic
import csv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
import sys

client = anthropic.Anthropic()

def categorize_expenses(expenses_text):
    message = client.messages.create(
        model="claude-opus-4-20250514",
        max_tokens=500,
        messages=[{"role": "user", "content": f"Categorize: {expenses_text}"}]
    )
    return message.content[0].text

def read_csv_file(filepath):
    expenses = []
    with open(filepath, 'r') as f:
        reader = csv.DictReader(f)
        for row in reader:
            expenses.append(row)
    return expenses

def format_expenses(expenses):
    return "\n".join([f"{e.get('description')} - £{e.get('amount')}" for e in expenses])

def export_to_excel(expenses, analysis):
    wb = Workbook()
    ws = wb.active
    ws.title = "Expenses"
    
    ws['A1'] = "EXPENSE REPORT"
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    ws['A2'] = f"Generated: {datetime.now().strftime('%d %B %Y')}"
    
    ws['A4'] = "Description"
    ws['B4'] = "Amount"
    for col in ['A', 'B']:
        ws[f'{col}4'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        ws[f'{col}4'].font = Font(bold=True)
    
    row = 5
    total = 0
    for exp in expenses:
        amount = float(exp.get('amount', 0))
        ws[f'A{row}'] = exp.get('description')
        ws[f'B{row}'] = amount
        ws[f'B{row}'].number_format = '£#,##0.00'
        total += amount
        row += 1
    
    ws[f'A{row}'] = "TOTAL"
    ws[f'B{row}'] = total
    ws[f'B{row}'].number_format = '£#,##0.00'
    ws[f'B{row}'].fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    
    wb.save("expense_report.xlsx")
    return "expense_report.xlsx"

if __name__ == "__main__":
    expenses = read_csv_file(sys.argv[1]) if len(sys.argv) > 1 else [{'description': 'Uber', 'amount': '12.50'}, {'description': 'Lunch', 'amount': '8.99'}]
    text = format_expenses(expenses)
    print("Analyzing...\n")
    result = categorize_expenses(text)
    print(result)
    file = export_to_excel(expenses, result)
    print(f"\n✓ Excel report created: {file}")
