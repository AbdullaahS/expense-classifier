import anthropic
import csv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from typing import List, Optional

class Expense:
    def __init__(self, description: str, amount: float) -> None:
        if not isinstance(description, str):
            raise TypeError(f"Description must be string")
        if not isinstance(amount, (int, float)):
            raise TypeError(f"Amount must be numeric")
        if amount <= 0:
            raise ValueError(f"Amount must be positive")
        self.description = description
        self.amount = float(amount)
    
    def __repr__(self) -> str:
        return f"Expense('{self.description}', £{self.amount})"

class ExpenseAnalyzer:
    def __init__(self) -> None:
        self.client = anthropic.Anthropic()
        self.expenses: List[Expense] = []
        self.analysis: Optional[str] = None
    
    def add_expense(self, description: str, amount: float) -> None:
        expense = Expense(description, amount)
        self.expenses.append(expense)
    
    def get_total(self) -> float:
        return sum(expense.amount for expense in self.expenses)
    
    def format_for_claude(self) -> str:
        if not self.expenses:
            raise ValueError("No expenses")
        return "\n".join([f"{e.description} - £{e.amount}" for e in self.expenses])
    
    def analyze(self) -> str:
        if not self.expenses:
            raise ValueError("No expenses")
        text = self.format_for_claude()
        message = self.client.messages.create(
            model="claude-opus-4-20250514",
            max_tokens=1000,
            messages=[{"role": "user", "content": f"Categorize: {text}"}]
        )
        self.analysis = message.content[0].text
        return self.analysis
    
    def export_to_excel(self, filename: str = "expense_report.xlsx") -> str:
        if not self.expenses:
            raise ValueError("No expenses")
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "EXPENSE REPORT"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"Generated: {datetime.now().strftime('%d %B %Y')}"
        ws['A4'] = "Description"
        ws['B4'] = "Amount"
        current_row = 5
        for exp in self.expenses:
            ws[f'A{current_row}'] = exp.description
            ws[f'B{current_row}'] = exp.amount
            ws[f'B{current_row}'].number_format = '£#,##0.00'
            current_row += 1
        ws[f'A{current_row}'] = "TOTAL"
        ws[f'B{current_row}'] = self.get_total()
        ws[f'B{current_row}'].number_format = '£#,##0.00'
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        wb.save(filename)
        return filename

if __name__ == "__main__":
    analyzer = ExpenseAnalyzer()
    analyzer.add_expense("Uber", 15.50)
    analyzer.add_expense("Lunch", 9.99)
    print(f"Total: £{analyzer.get_total()}")
